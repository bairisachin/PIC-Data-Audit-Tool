import os
import io
import re
import time
import openpyxl
import asyncio
import requests
import platform
from difflib import ndiff
from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import plotly.graph_objects as go
from plotly.graph_objs import *
import openpyxl.drawing.image as ImageDraw
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import PatternFill, Font, Border, Alignment, Side
from selenium.common.exceptions import WebDriverException
from DiffChecker import check_diff, get_diff
from OutlookEmail import EmailSender
# from ScrapeMatchTool import counter

class ExcelScraper:
    def __init__(self, emailTo, sourceFilePath, destinationFilePath, newFileName = "ScrapedSheet"):
        self.emailTo = emailTo

        self.sourceFilePath = sourceFilePath
        self.sourceWorkbook = openpyxl.load_workbook(self.sourceFilePath)
        self.sourceSheet = self.sourceWorkbook["Attributes"]
        # self.sourceSheet = self.sourceWorkbook["Animal Food"]
        self.sourceSheet = self.sourceWorkbook.active
        
        self.scrapeFilePath = destinationFilePath
        self.newFileName = f"{newFileName}.xlsx"
        self.fullFilePath = f"{self.scrapeFilePath}/{self.newFileName}"
        self.scrapeWorkbook = openpyxl.Workbook() # Create a new workbook
        self.scrapeSheet =  self.scrapeWorkbook.active
        self.scrapeSheet.title = "Audit Sheet" 
        # self.scrapeSheet = self.scrapeWorkbook.create_sheet("Scraped Sheet") # Create a new sheet as `Scraped Sheet``
        self.scrapeWorkbook.save(self.fullFilePath) # Save the file
        self.scrapeWorkbook.active = self.scrapeSheet # Set activate worksheet as scrapeSheet
        ColumnDimension(self.scrapeSheet, bestFit=True)

        self.headerRow = self.sourceSheet[5] # Get the header row
        self.headerMap = {} # Define a dictionary to hold the column indices for each header
        for cell in self.headerRow:
            self.headerMap[cell.value] = cell.column-1 # making the map 0-Based indexing

        # Defining the headers for scrape sheet
        self.scrapeSheetHeaders = ["UPC","GTIN","Model Number","Product Name", "Description", "Key Features 1", "Key Features 2", "Key Features 3", "Key Features 4", "Additional Features 1","Title","Desc","Key Feature","Items/Marketing images","Content Accuracy(%)","Audit status", "Remark", "URLs", "URLStatus"]
        self.scrapeSheetUnchangedCols = ["UPC", "GTIN", "Model Number", "URLs"]

        self.counter = 0 # This counts total products scraped
        self.stopFlag = True

        # Defining some varibles to keep track of matched values and there % ages
        self.total_products = self.sourceSheet.max_row - 5 # Remove the 1st 5 Headers
        self.titles_matched = 0
        self.description_matched = 0
        self.images_matched = 0
        self.key_feature1_matched = 0
        self.key_feature2_matched = 0
        self.key_feature3_matched = 0
        self.key_feature4_matched = 0
        self.additional_key_feature_matched = 0
        self.key_feature_matched = 0
        self.total_pass = 0
        self.total_fail = 0
        # Percentages variables
        self.title_percentage = 0
        self.description_percentage = 0
        self.images_percentage = 0
        self.key_feature_percentage = 0
        self.pass_percentage = 0
        self.fail_percentage = 0

        self.urlList = []
        self.dataList = []

    def main(self):
    #     # self.get_urlList_dataList()
        
        self.load_static_excel_content()
        self.load_dynamic_excel_content()
        self.compare_excel()

        self.build_walmart_dashboard()

        # Finally saving and closing workbook
        self.save_excel()
        self.close_excel()

        # Send email
        if self.emailTo is not None and self.emailTo != '':
            self.send_email(RECIPIENT=self.emailTo, FILE_PATH=self.fullFilePath)

    # def get_urlList_dataList(self):
    #     self.urlList = self.get_url()
    #     self.dataList = self.scrape_product_data()

    def build_walmart_dashboard(self):
        self.dashboardSheet = self.scrapeWorkbook.create_sheet(title="Dashboard") # Creating new sheet
        self.scrapeWorkbook.active = self.dashboardSheet # Setting it as active

        # Calculate percentages
        self.title_percentage = round(self.titles_matched / self.total_products * 100, 2)
        self.description_percentage = round(self.description_matched / self.total_products * 100, 2)
        self.images_percentage = round(self.images_matched / self.total_products * 100, 2)
        self.key_feature_percentage = round(self.key_feature_matched / self.total_products * 100, 2)
        # key_feature_percentage = (
        #     (self.key_feature1_matcheself.d + self.key_feature2_matched + self.key_feature3_matched +
        #     self.key_feature4_matched + self.additional_key_feature_matched) / (5 * self.total_products) * 100
        # )

        # Title
        self.dashboardSheet.merge_cells('B3:D4')
        cell = self.dashboardSheet.cell(row=3, column=2)
        cell.fill = PatternFill(start_color='4ade80', end_color='4ade80', fill_type='solid')
        cell.font = Font(bold=True)
        cell.value = "Wallmart Dashboard"
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

        cell = self.dashboardSheet.cell(row=5, column=3)
        cell.value = "Count"
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        cell = self.dashboardSheet.cell(row=5, column=4)
        cell.value = "%"
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

        # Fields 
        # TITLE MATCHED
        cell = self.dashboardSheet.cell(row=6, column=2)
        cell.value = "Titles Matched"
        # Set row/column width
        column_letter = cell.column_letter
        self.dashboardSheet.column_dimensions[column_letter].width = 25
        # TITLE MATCHED VALUE
        cell = self.dashboardSheet.cell(row=6, column=3)
        cell.value = f"{self.titles_matched}/{self.total_products}"
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        # TITLE MATCHED %
        cell = self.dashboardSheet.cell(row=6, column=4)
        cell.value = f"{self.title_percentage}%"
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        
        # DESC. MATCHED
        cell = self.dashboardSheet.cell(row=8, column=2)
        cell.value = "Description Matched"
        # DESC. MATCHED VALUE
        cell = self.dashboardSheet.cell(row=8, column=3)
        cell.value = f"{self.description_matched}/{self.total_products}"
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        # DESC. MATCHED %
        cell = self.dashboardSheet.cell(row=8, column=4)
        cell.value = f"{self.description_percentage}%"
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

        # KEY FEATURES 1 MATCHED
        cell = self.dashboardSheet.cell(row=10, column=2)
        cell.value = "Key Feature 1 Matched"
        # KEY FEATURES 1 MATCHED VALUE
        cell = self.dashboardSheet.cell(row=10, column=3)
        cell.value = f"{self.key_feature1_matched}/{self.total_products}"
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

        # KEY FEATURES 2 MATCHED
        cell = self.dashboardSheet.cell(row=12, column=2)
        cell.value = "Key Feature 2 Matched"
        # KEY FEATURES 2 MATCHED VALUE
        cell = self.dashboardSheet.cell(row=12, column=3)
        cell.value = f"{self.key_feature2_matched}/{self.total_products}"
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

        # KEY FEATURES 3 MATCHED
        cell = self.dashboardSheet.cell(row=14, column=2)
        cell.value = "Key Features 3 Matched"
        # KEY FEATURE 1 MATCHED VALUE
        cell = self.dashboardSheet.cell(row=14, column=3)
        cell.value = f"{self.key_feature3_matched}/{self.total_products}"
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

        # KEY FEATURES 4 MATCHED
        cell = self.dashboardSheet.cell(row=16, column=2)
        cell.value = "Key Features 4 Matched"
        # KEY FEATURES 4 MATCHED VALUE
        cell = self.dashboardSheet.cell(row=16, column=3)
        cell.value = f"{self.key_feature4_matched}/{self.total_products}"
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        
        # ADDITIONAL FEATURE MATCHED
        cell = self.dashboardSheet.cell(row=18, column=2)
        cell.value = "Additional Features Matched"
        # ADDITIONAL FEATURE MATCHED VALUE
        cell = self.dashboardSheet.cell(row=18, column=3)
        cell.value = f"{self.key_feature4_matched}/{self.total_products}"
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        
        # KEY FEATURE %
        self.dashboardSheet.merge_cells('D10:D18')
        cell = self.dashboardSheet.cell(row=10, column=4)
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        cell.value = f"{self.key_feature_percentage}%"

        # TOTAL PASSED 
        cell = self.dashboardSheet.cell(row=20, column=2)
        cell.value = "TOTAL PASSED"
        cell.fill = PatternFill(start_color='86efac', end_color='86efac', fill_type='solid')
        # TOTAL PASSED VALUE
        cell = self.dashboardSheet.cell(row=20, column=3)
        cell.value = f"{self.total_pass}/{self.total_products}"
        # COLOR GREEN
        cell.fill = PatternFill(start_color='86efac', end_color='86efac', fill_type='solid')
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        # TOTAL PASSED %
        cell = self.dashboardSheet.cell(row=20, column=4)
        cell.value = f"{(self.total_pass/self.total_products)*100}%"
        # COLOR GREEN
        cell.fill = PatternFill(start_color='86efac', end_color='86efac', fill_type='solid')
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        
        # TOTAL FAILED
        cell = self.dashboardSheet.cell(row=22, column=2)
        cell.value = "TOTAL FAILED"
        cell.fill = PatternFill(start_color='f87171', end_color='f87171', fill_type='solid')
        # TOTAL FAILED VALUE
        cell = self.dashboardSheet.cell(row=22, column=3)
        cell.value = f"{self.total_products - self.total_pass}/{self.total_products}"
        # COLOR RED
        cell.fill = PatternFill(start_color='f87171', end_color='f87171', fill_type='solid')
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        # TOTAL FAILED %
        cell = self.dashboardSheet.cell(row=22, column=4)
        cell.value = f"{((self.total_products - self.total_pass)/self.total_products)*100}%"
        # COLOR RED
        cell.fill = PatternFill(start_color='f87171', end_color='f87171', fill_type='solid')
        # Align center and wraptext
        cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        
        # Check if folder exists
        destination_folder = "Destination"
        graph_html_folder = os.path.join(destination_folder, "Graph_html")
        graph_png_folder = os.path.join(destination_folder, "Graph_png")

        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder)
            print("Destination folder created.")

        if not os.path.exists(graph_html_folder):
            os.makedirs(graph_html_folder)
            print("Graph_html folder created.")

        if not os.path.exists(graph_png_folder):
            os.makedirs(graph_png_folder)
            print("Graph_png folder created.")

        self.build_barplot()
        self.build_piechart()

        # Combine both fig obj
        with open('./Destination/Graph_html/Barplot_Piechart_Combined.html', 'a') as f:
            f.write(self.barplot_fig.to_html(full_html=False, include_plotlyjs='cdn'))
            f.write(self.piechart_fig.to_html(full_html=False, include_plotlyjs='cdn'))

    def build_barplot(self):
        # colors for bars
        red = 'rgba(248, 113, 113, 1)' # Failed
        green = 'rgba(74, 222, 128, 1)' # Passed
        yellow = 'rgba(250, 204, 21, 1)' # Titles
        teal = 'rgba(45, 212, 191, 1)' # Description
        purple = 'rgba(168, 85, 247, 1)' # key features


        # Setting BG to transparent
        layout = Layout(
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)'
        )
    
        # Create bar graph
        self.barplot_fig = go.Figure(layout=layout)
        self.barplot_fig.add_trace(go.Bar(
            x=[1],
            y=[self.titles_matched],
            name='Titles Matched',
            hovertext=f'{self.titles_matched} Passed',
            text=f'{self.titles_matched}',
            textposition='auto',
            marker=dict(color=yellow)
        ))
        self.barplot_fig.add_trace(go.Bar(
            x=[1],
            y=[self.total_products - self.titles_matched],
            showlegend=False,
            name='Titles Mismatched',
            hovertext=f'{self.total_products - self.titles_matched} Failed',
            text=f'{self.total_products - self.titles_matched}',
            textposition='auto',
            marker=dict(color=red)
        ))

        self.barplot_fig.add_trace(go.Bar(
            x=[2],
            y=[self.description_matched],
            name='Descriptions Matched',
            hovertext=f'{self.description_matched} Passed',
            text=f'{self.description_matched}',
            textposition='auto',
            marker=dict(color=teal)
        ))

        self.barplot_fig.add_trace(go.Bar(
            x=[2],
            y=[self.total_products - self.description_matched],
            showlegend=False,
            name='Descriptions Mismatched',
            hovertext=f'{self.total_products - self.description_matched} Failed',
            text=f'{self.total_products - self.description_matched}',
            textposition='auto',
            marker=dict(color=red)
        ))

        self.barplot_fig.add_trace(go.Bar(
            x=[3],
            y=[self.key_feature_matched],
            name='Key Features Matched',
            hovertext=f'{self.key_feature_matched} Passed',
            text=f'{self.key_feature_matched}',
            textposition='auto',
            marker=dict(color=purple)
        ))

        self.barplot_fig.add_trace(go.Bar(
            x=[3],
            y=[self.total_products - self.key_feature_matched],
            showlegend=False,
            name='Key Features Mismatched',
            hovertext=f'{self.total_products - self.key_feature_matched} Failed',
            text=f'{self.total_products - self.key_feature_matched}',
            textposition='auto',
            marker=dict(color=red)
        ))

        self.barplot_fig.add_trace(go.Bar(
            x=[4],
            y=[self.total_pass],
            name='Total Passed',
            hovertext=f'{self.total_pass} Passed',
            text=f'{self.total_pass}',
            textposition='auto',
            marker=dict(color=green)
        ))

        self.barplot_fig.add_trace(go.Bar(
            x=[4],
            y=[self.total_products - self.total_pass],
            showlegend=False,
            name='Total Failed',
            hovertext=f'{self.total_products - self.total_pass} Failed',
            text=f'{self.total_products - self.total_pass}',
            textposition='auto',
            marker=dict(color=red)
        ))


        # Set the layout
        self.barplot_fig.update_layout(
            title=f'<b>Walmart Product Matching Statistics of <span style="color:red;">{self.total_products}</span> Products</b>',
            yaxis_title="<b>Total No. of SKUs</b>",
            legend_title="<b>Product Properties</b>",
            # yaxis=dict(
            #     range=[0, self.total_products]  # Set the y-axis range to 0 to no. of products
            # ),
            xaxis=dict(
                tickmode='array',
                tickvals=[1, 2, 3, 4],
                ticktext=['Titles', 'Description', 'Key Features', 'Overall Passed']
            ),
            showlegend=True,
            barmode='stack',
            width=500, 
            height=500,
            # Adding a border to plot
        #     shapes=[go.layout.Shape(
        #     type='rect',
        #     xref='paper',
        #     yref='paper',
        #     x0=0,
        #     y0=-0.1,
        #     x1=1.01,
        #     y1=1.02,
        #     line={'width': 1, 'color': 'black'}
        # )]
        )

        # Save the fig as html
        self.barplot_fig.write_html("./Destination/Graph_html/Barplot.html")

        # Save the figure as a PNG image
        self.barplot_fig.write_image("./Destination/Graph_png/Barplot.png", format="png")
        print("Barplot Complete!")

        # Add the graph as a picture to the worksheet
        img = ImageDraw.Image("./Destination/Graph_png/Barplot.png")
        img.anchor = self.dashboardSheet.cell(row=3, column=6).coordinate  # Cell F3
        img.width = 400
        img.height = 400
        self.dashboardSheet.add_image(img)

    def build_piechart(self):
        # Calculate percentages
        self.pass_percentage = round(self.total_pass / self.total_products * 100, 2)
        self.total_fail = self.total_products - self.total_pass
        self.fail_percentage = round(self.total_fail / self.total_products * 100, 2) # Round off till 2 digit
        # self.fail_percentage = 100 - self.pass_percentage
        # Define the colors for pass and fail
        red = 'rgba(248, 113, 113, 1)' # Failed
        green = 'rgba(74, 222, 128, 1)' # Passed
        colors = [green, red]
        lables = ['SKUs >= 75%', 'SKUs < 75%']
        values = [self.total_pass,self.total_fail]

        # Setting BG to transparent
        layout = Layout(
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            legend_title='<b>Content Accuracy</b>'
        )

        # Create pie chart
        self.piechart_fig = go.Figure(data=go.Pie(
                        labels=lables, 
                        # values=[self.pass_percentage, self.fail_percentage],
                        values=values,
                        marker=dict(colors=colors),
                        pull=[0, 0.1], # Explote the 2nd slice(Fail) by 0.1
                        # textinfo=f'value+percent',
                        text=[f'{self.total_pass} SKUs', f'{self.total_fail} SKUs']
                        # text=[f'{self.total_pass} SKUs <br>{self.pass_percentage}', f'{self.total_fail} SKUs <br>{self.fail_percentage}']
                        ), layout=layout)

        # Set the layout
        self.piechart_fig.update_layout(
            title='<b> Product Pass/Fail Ratio </b>',
            width=500,
            height=500,
            # Adding a border to the plot
        #     shapes=[go.layout.Shape(
        #     type='rect',
        #     xref='paper',
        #     yref='paper',
        #     x0=0,
        #     y0=-0.1,
        #     x1=1.01,
        #     y1=1.02,
        #     line={'width': 1, 'color': 'black'}
        # )]
        )
        
        # Save the fig to html
        self.piechart_fig.write_html("./Destination/Graph_html/Piechart.html")

        # Save the figure as a PNG image
        self.piechart_fig.write_image("./Destination/Graph_png/Piechart.png", format="png")
        print("Piechart Complete!")
        
        # Add the graph as a picture to the worksheet
        img = ImageDraw.Image("./Destination/Graph_png/Piechart.png")
        img.anchor = self.dashboardSheet.cell(row=3, column=13).coordinate  # Cell M3
        img.width = 400
        img.height = 400
        self.dashboardSheet.add_image(img)

    def compare_excel(self):
        # Saving the changes of Audit Sheet before making the Republish Sheet as active
        self.save_excel()
        # Creating new sheet for republish data
        self.republishSheet = self.scrapeWorkbook.create_sheet(title="Republish Sheet")
        self.scrapeWorkbook.active = self.republishSheet

        # Copy/paste the entire header from sourceExel to scrapeExcel
        for row in self.sourceSheet.iter_rows(min_row=1, max_row=5): # Appending the first 5 header rows to scrapeSheet            
            rowValues = [cel.value for cel in row]
            self.republishSheet.append(rowValues)

        # Apply styles to row 5
        for cell in self.republishSheet[5]:
            # Styling the header 
            cell.fill = PatternFill(start_color='4ade80', end_color='4ade80', fill_type='solid')
            cell.font = Font(bold=True)
            # Set row/column width
            column_letter = cell.column_letter
            self.republishSheet.row_dimensions[cell.row].height = 45
            self.republishSheet.column_dimensions[column_letter].width = 15
            # Align center and wraptext
            cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
            
        totalPassedSKUs = 0

        thin = Side(border_style="thin", color='FF0000')
        # Comparing
        for row in self.sourceSheet.iter_rows(min_row=6):
        
            currRowIdx = row[0].row - 4 # NOTE: this currRowIdx is for scrapeSheet and not for sourceSheet

            # Flags to check eachone
            correctProdNamePercent = 0
            correctDescPercent = 0
            # keyFeatureFlag = True
            correctKeyFeaturePercent = 0

            # vars to store difference after check_diff func is applied
            titleDiff = ""
            descDiff = ""
            kf1Diff = ""
            kf2Diff = ""
            kf3Diff = ""
            kf4Diff = ""
            kf5Diff = ""
            remark = ""
        
            if check_diff(row[self.headerMap["Product Name"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Product Name"]).value):
                self.titles_matched+=1
                correctProdNamePercent = 25
            else:
                # Red border added
                self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Product Name"]).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                titleDiff = get_diff(row[self.headerMap["Product Name"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Product Name"]).value)
                # if self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Product Name"]).value != "":
                #     # a, b, uncommon = self.get_difference(row[self.headerMap["Product Name"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Product Name"]).value)
                #     cellText = f"Title: {titleDiff} \nDescription: "" \nKeyFeatures: """
                #     self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Remark"]).value = cellText
                #     self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Remark"]).alignment = Alignment(horizontal='center', vertical='center')
                # else:
                #     self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Remark"]).value = "-"
                #     self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Remark"]).alignment = Alignment(horizontal='center', vertical='center')
                    

            if check_diff(row[self.headerMap["Description"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Description"]).value):
                self.description_matched+=1
                correctDescPercent = 25
                pass
            else:
                # Red border added
                self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Description"]).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                descDiff = get_diff(row[self.headerMap["Description"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Description"]).value)

            if check_diff(row[self.headerMap["Key Features 1"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 1"]).value):
                self.key_feature1_matched+=1
                correctKeyFeaturePercent+=5 # Add 5% for each correct Key feature
                pass
            else:
                # Red border added
                self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 1"]).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                kf1Diff = get_diff(row[self.headerMap["Key Features 1"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 1"]).value)

            if check_diff(row[self.headerMap["Key Features 2"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 2"]).value):
                self.key_feature2_matched+=1
                correctKeyFeaturePercent+=5
                pass
            else:
                # Red border added
                self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 2"]).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                kf2Diff = get_diff(row[self.headerMap["Key Features 2"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 2"]).value)

            if check_diff(row[self.headerMap["Key Features 3"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 3"]).value):
                self.key_feature3_matched+=1
                correctKeyFeaturePercent+=5
                pass
            else:
                # Red border added
                self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 3"]).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                kf3Diff = get_diff(row[self.headerMap["Key Features 3"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 3"]).value)

            if check_diff(row[self.headerMap["Key Features 4"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 4"]).value):
                self.key_feature4_matched+=1
                correctKeyFeaturePercent+=5
            else:
                # Red border added
                self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 4"]).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                k4Diff = get_diff(row[self.headerMap["Key Features 4"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Features 4"]).value)

            if check_diff(row[self.headerMap["Additional Features 1"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Additional Features 1"]).value):
                self.additional_key_feature_matched+=1
                correctKeyFeaturePercent+=5 
            else:
                # Red border added
                self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Additional Features 1"]).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                kf5Diff = get_diff(row[self.headerMap["Additional Features 1"]].value, self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Additional Features 1"]).value)

            
            column_letter = self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Remark"]).column_letter
            if (correctProdNamePercent + correctDescPercent + correctKeyFeaturePercent >= 75): # >= 75% is the passing condition
                totalPassedSKUs+=1

                # update the Remark col
                self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Remark"]).value = "-"
                self.scrapeSheet.column_dimensions[column_letter].width = 40
                self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Remark"]).alignment = Alignment(horizontal='center', vertical='center')
            else:
                currRow = self.sourceSheet[row[0].row] # Get the current row by row[0].row
                rowValues = []
                
                for cell in currRow: # Copy entire row from sourceSheet to republishSheet
                    # if cell.hyperlink is not None and cell.column == self.headerMap["URLs"]:

                    #     rowValues.append(cell.hyperlink.target)
                    #     continue
                    if cell.value is not None and cell.column == self.headerMap["URLs"] and (cell.hyperlink is not None and cell.hyperlink.target is not None):
                        # print(cell.value)
                        rowValues.append(cell.hyperlink.target)
                        continue
                    rowValues.append(cell.value)
                    
                self.republishSheet.append(rowValues)

                # update the Remark col
                if titleDiff != '':
                    remark += f"Title: {titleDiff}\n"
                if descDiff != '':
                    remark += f"\nDescription: {descDiff}\n"
                if kf1Diff != '' or kf2Diff != '' or kf3Diff != '' or kf4Diff != '' or kf5Diff != '':
                    remark += f"\nKF: [\n1] {kf1Diff}\n\n2] {kf2Diff},\n\n3] {kf3Diff},\n\n4] {kf4Diff},\n\n5]{kf5Diff}\n]"
                
                if remark == "":
                    self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Remark"]).value = '-'
                    self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Remark"]).alignment = Alignment(horizontal='center', vertical='center')
                    self.scrapeSheet.column_dimensions[column_letter].width = 40
                else:
                    # remark = f"Title: {titleDiff}\n\nDescription: {descDiff}\n\nKF: [{kf1Diff},\n {kf2Diff},\n {kf3Diff},\n {kf4Diff}, {kf5Diff}]'"
                    self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Remark"]).value = remark
                    self.scrapeSheet.column_dimensions[column_letter].width = 40
                    # self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Remark"]).alignment = Alignment(horizontal='center', vertical='center')
                    
            if correctKeyFeaturePercent == 25:
                self.key_feature_matched+=1

            self.calculate_accuracy(currRowIdx=currRowIdx, correctProdNamePercent=correctProdNamePercent, correctDescPercent=correctDescPercent, correctKeyFeaturePercent=correctKeyFeaturePercent)
        
        self.save_excel()
    
    def calculate_accuracy(self, currRowIdx, correctProdNamePercent, correctDescPercent, correctKeyFeaturePercent, correctImgPercent = 0):
        
        # Product title
        # Align center and wraptext
        self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Title"]).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        if correctProdNamePercent:
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Title"]).value = f"{correctProdNamePercent}%"
        else:
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Title"]).value = f"{correctProdNamePercent}%"
        
        # Description
        self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Desc"]).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        if correctDescPercent:
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Desc"]).value = f"{correctDescPercent}%"
        else:
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Desc"]).value = f"{correctDescPercent}%"

        # Key feature
        self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Feature"]).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        if correctKeyFeaturePercent:
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Feature"]).value = f"{correctKeyFeaturePercent}%"
        else:
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Key Feature"]).value = f"{correctKeyFeaturePercent}%"

        # Image
        self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Items/Marketing images"]).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        # both incorrect because we are not working on images
        if correctImgPercent:
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Items/Marketing images"]).value = f"{correctImgPercent}%"
        else:
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Items/Marketing images"]).value = f"{correctImgPercent}%"

        # Content accuracy
        self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Content Accuracy(%)"]).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Content Accuracy(%)"]).value = f'{correctProdNamePercent + correctDescPercent + correctKeyFeaturePercent}%'

        # Audit status
        self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Audit status"]).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        if (correctProdNamePercent + correctDescPercent + correctKeyFeaturePercent) >= 75 : # Pass
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Audit status"]).value = "Pass"
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Audit status"]).fill = PatternFill(start_color='86efac', end_color='86efac', fill_type='solid')
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Audit status"]).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
            self.total_pass+=1
        else: # Fail
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Audit status"]).value = "Fail"
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Audit status"]).fill = PatternFill(start_color='f87171', end_color='f87171', fill_type='solid')
            self.scrapeSheet.cell(row=currRowIdx, column=self.scrapeSheetHeaderMap["Audit status"]).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

    def load_static_excel_content(self):
        staticColumns = []
        # staticColumns = ["UPC", "GTIN", "Model Number", "URLs"]
        
        staticRows = [self.scrapeSheetUnchangedCols]

        # Loop through each row in sourceSheet starting from the 6th row
        for row in self.sourceSheet.iter_rows(min_row=6):
            # Create an empty list to store values from included columns
            rowValue = []
            # Loop through each cell in the row and add the value to the rowValue list if the column is included
            for idx, cell in enumerate(row):
                if self.sourceSheet.cell(row=5, column=idx+1).value in self.scrapeSheetUnchangedCols:
                    # if self.sourceSheet.cell(row=5, column=idx+1).value == "URLs":
                    #     # print(cell.value)
                    #     rowValue.append(self.urlList[(cell.row) - 6])
                    #     continue
                    if cell.value is not None and (self.sourceSheet.cell(row=5, column=idx+1).value == "URLs" and (cell.hyperlink is not None and cell.hyperlink.target is not None)):
                        # print(cell.value)
                        rowValue.append(cell.hyperlink.target)
                        continue
                    rowValue.append(cell.value)
            
            staticRows.append(rowValue)
        
        # These are all the cloumns of scrapedSheet
        self.scrapeSheet.append(self.scrapeSheetHeaders)

        headerRow = self.scrapeSheet[1] # Get the header row
        self.scrapeSheetHeaderMap = {} # Define a dictionary to hold the column indices for each header
        for cell in headerRow: 
            # Creating scrapeSheetHeaderMap
            self.scrapeSheetHeaderMap[cell.value] = cell.column
            
            # Styling the header #22c55e
            cell.fill = PatternFill(start_color='4ade80', end_color='4ade80', fill_type='solid')
            cell.font = Font(bold=True)

            # Set row/column width
            column_letter = cell.column_letter
            self.scrapeSheet.row_dimensions[cell.row].height = 30
            self.scrapeSheet.column_dimensions[column_letter].width = 15

            # Align center and wraptext
            cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

        # Update the values in respective columns for each row in the scrapeSheet
        rowNo = 2
        for idx, row in enumerate(staticRows):
            if idx == 0: # Skip the first row
                continue
            # Add the static row to the scrapeSheet
            self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["UPC"]).value = row[0]
            self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["GTIN"]).value = row[1]
            self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["Model Number"]).value = row[2]
            self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["URLs"]).value = row[3]
            if self.urlList[idx-1] == '#PAGE NOT FOUND' or self.urlList[idx-1] == '#URL NOT FOUND':
                self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["URLStatus"]).value = self.urlList[idx-1]
                self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["URLStatus"]).alignment = Alignment(horizontal='center', vertical='center')
            else:
                self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["URLStatus"]).value = '-'
                self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["URLStatus"]).alignment = Alignment(horizontal='center', vertical='center')
            rowNo+=1
    
    def load_dynamic_excel_content(self):

        rowNo = 2
        for idx, row in enumerate(self.dataList): 
            if idx == 0: # Skip the first row
                continue
            # Add the dynamic row values to the scrapeSheet
            self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["Product Name"]).value = row[0]
            self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["Description"]).value = row[1] 
            self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["Key Features 1"]).value = row[2]
            self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["Key Features 2"]).value = row[3]
            self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["Key Features 3"]).value = row[4]
            self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["Key Features 4"]).value = row[5]
            self.scrapeSheet.cell(row=rowNo, column=self.scrapeSheetHeaderMap["Additional Features 1"]).value = row[6]
            rowNo+=1

    async def scrape_product_data(self):
        webdriver_manager = ChromeDriverManager(path='./chromedrivers')

        # # Check if the installed ChromeDriver version is up-to-date or not
        # current_version = platform.release()  # Get the current version of Chrome
        # latest_version = webdriver_manager.install()
        # if current_version != latest_version:
            # print("Updating ChromeDriver...")
            # webdriver_manager.install()

        # Get the path of the installed ChromeDriver
        webdriver_path = webdriver_manager.install()
                
        # Use ChromeDriverManager to download and install ChromeDriver once
        # webdriver_manager.install()

        # _driverpath = r'C:\\Users\\G688301\\OneDrive - General Mills\\Desktop\\WebScrapeUpdated\\chromedriver_win32\\chomedriver_v114.exe'
        
        dataList = [['Product Name','Description','Key Features 1','Key Features 2','Key Features 3','Key Features 4','Additional Features 1']]

        idx = 0

        while (idx < self.total_products):

            # if self.counter in [57, 58,59,100]:
            #     flag = True
            # driver = webdriver.Chrome(executable_path = _driverpath, options=_options)
            # driver = webdriver.Chrome(executable_path = _driverpath)
            try:
                if self.urlList[idx] != "#URL NOT FOUND":
                    # time.sleep(3)

                    # Use ChromeDriverManager to automatically download and install the correct ChromeDriver version
                    driver = self.initialize_driver(webdriver_path) 
                    # try:
                    driver.get(self.urlList[idx])

                    if self.is_session_expired(driver): # NOTE : Checking if the driver session has expired or not
                        print("Session expired. Handling session expiration...")
                        driver.quit()
                        driver = self.initialize_driver()
                        # complete the unfinished get request with new driver instance
                        driver.get(self.urlList[idx])

                    # except WebDriverException:
                    #     print("page down")
                    #     self.urlList[idx] = '#PAGE DOWN'
                        
                    html = driver.page_source # Retrieve the HTML source code

                    soup = BeautifulSoup(html, 'html.parser') # Convert the HTML source code to plain text

                    # self.save_screenshot(driver, self.counter)
                    # self.save_html(soup, self.counter)
                    
                    h1Text, spanText = self.check_page_not_found(soup)
                    if h1Text == "Uh-oh..." or spanText == "This page could not be found.":
                        dataList.append(["","","","","","",""]) # Page not found so fill empty values for that row
                        self.urlList[idx] = '#PAGE NOT FOUND'
                    else:
                        keyFeatures = self.get_keyfeatures(soup)
                        # dataList.append([self.get_product_name(soup),self.get_description(soup),keyFeatures[0],keyFeatures[1],keyFeatures[2],keyFeatures[3],keyFeatures[4]])
                        dataList.append([self.get_product_name(soup),self.get_description(soup), *keyFeatures])
                else :
                    # Update empty list when url not found
                    dataList.append(["","","","","","",""])
                
                if self.counter == self.total_products - 1:
                    self.stopFlag = False

                self.counter+=1
                await asyncio.sleep(0.3)
                print(f"{self.counter}/{self.total_products}")
                idx += 1

                if idx % 10 == 0: # After every 10 SKUs do the cleanup
                    # Clear cookies, local storage, or perform other necessary cleanup
                    driver.delete_all_cookies()
                    driver.execute_script("window.localStorage.clear();")

                # driver.quit()
            except Exception as ex:
                print("An exception occurred:", str(ex))
                self.stopFlag = False

                # idx-=1
                # dataList.pop() # remove the last entry
                # print("Restarting Driver...")
                # driver.quit()
                # time.sleep(5)
                # break

        # Close the driver
        driver.quit() 
                
        # return dataList
        self.dataList = dataList

    def get_url(self):
        productUrls = []
        for row in self.sourceSheet.iter_rows(min_row=6):
            rowNo = row[0].row
            urlCell = self.sourceSheet.cell(row=rowNo, column=self.headerMap["URLs"]+1)
            if urlCell.hyperlink is not None:
                productUrls.append(urlCell.hyperlink.target)
            elif urlCell.value is not None and 'www.walmart.com/ip/' in urlCell.value: #urlCell.value.__contains__('https://www.walmart.com/ip/')
                productUrls.append(urlCell.value)
            else:
                productUrls.append("#URL NOT FOUND")
        # return productUrls
        self.urlList = productUrls

    def check_page_not_found(self, soup):
        try:
            #  w_U9_0 w_U0S3 w_QcqU mb6 tc
            page_not_found_h1 = soup.find('h1',{'class': 'w_97UH w_O_Ib w_fwGT mb5'}).text.strip()
            try:
                page_not_found_span = soup.find('span',{'class': 'w_U9_0 w_U0S3 w_QcqU mb6 tc'}).text.strip()
            except:
                page_not_found_span = soup.find('span',{'class': 'w_97UH w_O_Ib w_U9_0 w_U0S3 w_QcqU mb6 tc mb5'}).text.strip()
                
        except:
            page_not_found_h1 = ""
            page_not_found_span = ""

        return page_not_found_h1, page_not_found_span

    def get_product_name(self, soup):
        try:
            productName = soup.find('h1', {'itemprop': 'name'}).text.strip()
        except:
            productName = ""
        return productName

    def get_description(self, soup):
        try:
            description = soup.find_all('div', {'class': 'dangerous-html mb3'})[0].text.strip()
        except:
            description = ""
        return description

    def get_keyfeatures(self, soup):
        try:
            divTag = soup.find_all('div', {'class': 'dangerous-html mb3'})[1]
            liTags = divTag.find_all('li') # Find all li tags inside the div tag
            keyFeatures = [li.text.strip() for li in liTags] # Extract the text inside the li 
        except:
            keyFeatures = ["","","","",""]

        return keyFeatures
        
    # UTILITY FUNCTIONs
    def save_screenshot(self, driver, cnt):
        output_folder = 'Screenshots'
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        driver.save_screenshot(f"./Screenshots/screenshot{cnt}.png")
        driver.get_screenshot_as_file(f"./Screenshots/screenshot{cnt}.png")
        print(driver.title)

    def save_html(self, soup, cnt):
        output_folder = 'HTML'
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Create a new HTML file and write the prettified data to it
        output_file = './HTML/index.html'
        output_file = f"./HTML/index{cnt}.html"
        with io.open(output_file, 'w', encoding='utf-8') as file:
            file.write(soup.prettify())

    def normalize_sentence(self, sentence):
        pattern = r'[^a-zA-Z0-9\s]'
        # pattern2 = r'[^A-Za-z0-9]+'
        # pattern3 = r'[^a-zA-Z0-9\s]+'
        # using regex to remove specialCharacters
        if sentence :
            new_sentence = ((re.sub(pattern, '', sentence)).lower()).strip()
            # new_sentence = ((re.sub(pattern3, '', sentence)).lower()).strip()
            return new_sentence
        return sentence

    # def check_diff(self, string1, string2):
    #     # Split both strings into list items
    #     string1 = self.normalize_sentence(string1).split() # Remove Special Chars + Lowercase + strip extra space
    #     string2 = self.normalize_sentence(string2).split()

    #     A = set(string1) # Store all string1 list items in set A
    #     B = set(string2) # Store all string2 list items in set B
        
    #     str_diff = A.symmetric_difference(B)
    #     return str_diff
    
    # def get_difference(self, string1, string2):
    #     # Split both strings into list items
    #     string1 = self.normalize_sentence(string1).split() # Remove Special Chars + Lowercase + strip extra space
    #     string2 = self.normalize_sentence(string2).split()

    #     A = set(string1) # Store all string1 list items in set A
    #     B = set(string2) # Store all string2 list items in set B
        
    #     uniqueA = A - B
    #     uniqueB = B - A
    #     uncommon = A.symmetric_difference(B)

        return uniqueA, uniqueB, uncommon

    def save_excel(self):
        '''
        We are only saving the scrapeWorkbook because this contains the excelSheets which 
        we are updating and sourceWorkbook is only used to read content
        '''
        self.scrapeWorkbook.save(f"{self.scrapeFilePath}/{self.newFileName}")

    def initialize_driver(self, webdriver_path):
        # driver configuration for headless mode
        _userAgent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36', 'Accept-Language': 'en-US, en;q=0.5"

        _options = webdriver.ChromeOptions()
        # _options.headless = True
        # _options.add_argument(f'--user-agent={_userAgent}')
        _options.add_argument("--window-size=320,180")
        _options.add_argument('--disable-gpu')

        # Initialize and return a new driver instance
        return webdriver.Chrome(executable_path=webdriver_path, options=_options)

    def is_session_expired(self, driver):
        # Check the HTTP status code
        response = requests.get(driver.current_url)
        if response.status_code == 401:
            return True  # Session expired

        # Check for the presence of a specific cookie
        cookies = driver.get_cookies()
        if any(cookie["name"] == "session_expired" for cookie in cookies):
            return True  # Session expired

        return False  # Session not expired

    def close_excel(self):
        # Closing both files to save memory
        self.sourceWorkbook.close()
        self.scrapeWorkbook.close()

    def send_email(self, RECIPIENT, FILE_PATH):
        email = EmailSender()
        email.send_email(EMAIL_TO=RECIPIENT, FILE_PATH=FILE_PATH)
        # email.send_emails(recipient=RECIPIENT, attachment_path=FILE_PATH)
